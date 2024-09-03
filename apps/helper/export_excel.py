import pandas as pd
from django.http import HttpResponse
from openpyxl.styles import PatternFill

def export_to_excel(queryset, filename='export.xlsx', columns=None, total_cost=None):
    """
    Export a queryset to an Excel file with a total cost row at the end.

    Args:
        queryset: The Django queryset to be exported.
        filename (str): The name of the resulting Excel file.
        columns (dict): A dictionary where keys are model field names and values are the column names in the Excel file.
        total_cost (float): The total cost to add as a final row in the Excel file.
    """
    # Convert queryset to a list of dictionaries
    data = list(queryset.values(*columns.keys())) if columns else list(queryset.values())

    # Rename columns if specified
    if columns:
        data = [{columns.get(k, k): v for k, v in item.items()} for item in data]
    
    # Create DataFrame
    df = pd.DataFrame(data)

    # Add the total row if total_cost is provided
    if total_cost is not None:
        # Create a new row with 'TOTAL' in the first column and total_cost in the last column
        total_row = {key: '' for key in df.columns}  # Initialize all columns with empty strings
        total_row[df.columns[0]] = 'TOTAL'  # Set the first column to 'TOTAL'
        total_row[df.columns[-1]] = total_cost  # Set the last column to total_cost
        
        # Convert the total_row dictionary to a DataFrame and transpose it to make it a row
        total_row_df = pd.DataFrame([total_row])

        # Concatenate the total_row_df with the original df
        df = pd.concat([df, total_row_df], ignore_index=True)

    # Create HTTP response with Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Write DataFrame to Excel file in the response
    with pd.ExcelWriter(response, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Define styles
        header_fill = PatternFill(start_color='87CEEB', end_color='87CEEB', fill_type='solid')  # Light Blue
        total_fill = PatternFill(start_color='98FB98', end_color='98FB98', fill_type='solid')  # Light Green

        # Apply header style to the first row (headers)
        for cell in worksheet[1]:
            cell.fill = header_fill

        # Apply total style to the last row (total row)
        for cell in worksheet.iter_rows(min_row=worksheet.max_row, max_row=worksheet.max_row):
            for c in cell:
                c.fill = total_fill

    return response
